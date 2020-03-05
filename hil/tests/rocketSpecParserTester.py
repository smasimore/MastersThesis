import sys
sys.path.insert(1, '../src') # Inserts src folder so we can import the code to test
from rocketSpecParser import RocketSpecParser, DEFAULT_WORKBOOK_NAME, DEFAULT_SHEET_NAME, DEFAULT_EXTENSION
from openpyxl import Workbook

# Creates a fake test workbook to test that the parser can properly
# parse out the info and store it in memory correctly
class RocketSpecParserTester():

    ### Start of Static Variables
    
    headers = ["Common Name", "Sensor Name", "Sensor type", "Communication protocol", "Pin"] # Headers for the fake test workbook
    testSensors = [["ParachuteSensor", "esp32", "analog", "None", "2"],                      # fake sensors to populate the test workbook
                   ["AltitudeSensor", "hc05", "analog", "None", "5"],
                   ["IMU", 'MPU6050', "None", "i2c", "7"]]
    
    ### End of static Variables

    ### Instance Methods

    # Runs through multiple sensors checking that the spec parser reads out their specs correctly
    def testSpecParser(self):
        print('Initialzing test spec sheet')
        self.initTestWorkbook()
        self.parser = RocketSpecParser()
        print('Spec sheet initialized, moving to parsing test..')
        for row in RocketSpecParserTester.testSensors:
            sensorObject = self.parser.getSensorInfo(row[0])
            sensorInfo = sensorObject.getSensorSpecs()
            comparison = [str(elem1) == elem2 for elem1, elem2 in list(zip(sensorInfo, row))]
            assert(all(comparison) == True) # Checking that all elements are equal
        print('No errors thrown during testing, TEST PASSED')

    # initializes a fake workbook for the spec parser to read from for testing
    def initTestWorkbook(self): 
        workbook = Workbook()
        sheet = workbook.create_sheet(DEFAULT_SHEET_NAME, 0)
        sheet.append(RocketSpecParserTester.headers)
        for row in RocketSpecParserTester.testSensors:
            sheet.append(row)
        workbook.save(filename= DEFAULT_WORKBOOK_NAME + DEFAULT_EXTENSION)
        
    ### End of instance methods

if __name__ == "__main__":
    tester = RocketSpecParserTester()
    tester.testSpecParser()
