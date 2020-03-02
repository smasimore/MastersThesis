import socket, threading, struct, os
from openpyxl import Workbook, load_workbook
from datetime import datetime

### Global variable define statements
MAX_MESSAGE_SIZE    = 1024                                                          # Max number of bytes a packet can be
DEFAULT_TEMPLATE = "fff?iLlhd"                                                      # Template of what data types are in the DataVector, will be loaded from a config file later
DEFAULT_LOG_DIRECTORY_PATH = "SimulationLogs-"
DEFAULT_LOG_NAME = "SimulationLog"
DEFAULT_COMMAND_SHEET_NAME = "Command History"
DEFAULT_TELEMETRY_SHEET_NAME = "DataVector Logs" 
DEFAULT_EXTENSION = '.xlsx'                                                         # All logs will be excel spreadsheets
DEFAULT_IP = "0.0.0.0"                                                              # 0.0.0.0 means accept all incoming connections                
DEFAULT_LOGGING_PORT = 8080                                                         # The default port the telemetry packets will come in on
### End of Global Variable Defines

# SimulationLogger is responsible for keeping a time stamped log of
# everything going on in the simulation that it was insantiated in.
# It has an open socket accepting ethernet packets from the flight
# computer and will store the most recent DataVector in memory.
# The logger inherits from the Thread class in order for it to have
# it's own threading functionality. Once initialized, the start()
# method can be called to start the instantiated object running it's
# run method as a seperate thread apart from the main simulation.
# This thread sits on the socket and blocks waiting on telemetry
# packets from the flight computer. This allows for the logger to
# block on telemetry packets without blocking the entire simulation.
# While the logger is handling incoming telemetry packets and logs
# them the simulation is still able to call the instance methods of
# the logger. So even if the logger is blocked the simulation is
# still able to come in and read the most up to date DataVector.
# To set everything up correctly though the super constructor of
# the parent class, Thread in this case, must also be called during
# initialization of the logger. 
class SimulationLogger(threading.Thread):

    ### Start of instance methods ###
    
    def __init__(self, template = DEFAULT_TEMPLATE, IP = DEFAULT_IP, port = DEFAULT_LOGGING_PORT): 
        self.sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)                # UDP socket  
        self.sock.bind((IP, port))                                                  # The logger acts as a server waiting for the flight computer to connect to it
        
        self.template = template                                                    # The template is a description of what data types the Data Vector consists of
        self.currentDataVector = []                                                 # Stores the last received Data Vector from the flight computer
        self._initializeSimulationLog()

        self.lock = threading.Lock()                                                # A lock is used when updating the current data vector to avoid race conditions
        super(SimulationLogger, self).__init__()                                    # Calls the specified parent method, SimulationLogger inherits from the Thread class, so Thread's constructor is called to set up everything right

    # All Threads need a run method that is called when Thread.start() is called
    def run(self):
        while True:
            dataVector = self._receiveDataVector()
            self._updateCurrentDataVector(dataVector)
            self._logDataVector(dataVector)

    # Creates a directory and excel workbook to store simulation logs
    def _initializeSimulationLog(self):
        self._initializeLogDirectory()                                              
        timeStamp = str(datetime.time(datetime.now())).split('.')[0]
        timeStamp = timeStamp.replace(':', '-', 2)                                  # Replacing invalid naming characters
        self.logName = DEFAULT_LOG_NAME + '-' + timeStamp + DEFAULT_EXTENSION       # The workbook name for this simulation is time stamped with the current hh-mm--ss time
        self.workbook = Workbook()
        self.workbook.create_sheet(DEFAULT_COMMAND_SHEET_NAME, 0)                   # Stores commands issued by the simulato    
        self.workbook.create_sheet(DEFAULT_TELEMETRY_SHEET_NAME, 0)                 # Stores all received Data Vectors from the flight computer

    # Creates a directory time stamped with todays date to store the logs
    def _initializeLogDirectory(self):                                               
        date = str(datetime.date(datetime.today()))
        self.dirName = DEFAULT_LOG_DIRECTORY_PATH + date
        if not os.path.isdir(self.dirName):                                         # Don't make the dir if it already exists
            os.mkdir(self.dirName)

    # Receives and unpacks a byte array from an incoming UDP packet
    def _receiveDataVector(self):
        byteData, _ = self.sock.recvfrom(MAX_MESSAGE_SIZE)                          
        dataVector = list(struct.unpack(self.template, byteData))                   # Unpacks the byteData into an array of data types detailed in the template
        return dataVector

    # Updates the Datavector saved in memory with the latest received vector
    def _updateCurrentDataVector(self, dataVector):
        self.lock.acquire()
        self.currentDataVector = dataVector                                         
        self.lock.release() 

    # Logs a time stamped DataVector to the DataVector sheet
    def _logDataVector(self, dataVector):
        self._logToSheet(DEFAULT_TELEMETRY_SHEET_NAME, dataVector)

    # Thread safe way to read the current DataVector stored in memory
    def readDataVector(self):
        self.lock.acquire()
        dataVector = self.currentDataVector
        self.lock.release()
        return dataVector

    # TODO: figure out DataVector elem names and implement name specific indexing
    def readDataVectorElement(self):
        pass

    # Logs a time stamped command to the command sheet
    def logCommandInput(self, command):
        newRow = [command]
        self._logToSheet(DEFAULT_COMMAND_SHEET_NAME, newRow)      

    # Adds a time stamp to the input row and logs it to the desired sheet
    def _logToSheet(self, sheetName, row):
        sheet = self.workbook[sheetName]
        timeStamp = str(datetime.time(datetime.now()))                              # Time stamp each log entry for a time record of the simulation
        row.append(timeStamp)        
        sheet.append(row)
        self.workbook.save(filename = os.path.join(self.dirName, self.logName))          

    # Returns the file path to the log workbook
    def getLogPath(self):
        return os.path.join(self.dirName, self.logName)

    ### End of Instance Methods ###
    
if __name__ == "__main__":
    logger = SimulationLogger()
    logger.start()

    dataArray = [3.0, 4.0, 5.0, 48, 6]                                              # Example Data Vector to send
    sock = socket.socket()
    sock.connect(("127.0.0.1", 8080))
    print("Connected to server")

    ## Basic test code where the user can manually watch the unpacking of a specified Data Vector
    while True:
        input("Press enter to send next message")
        sock.send(struct.pack(DEFAULT_TEMPLATE, *dataArray))
        input("Message sent, press enter to continue")
        print("Message received by logger: {0}".format(logger.readDataVector()))
        logger.logCommandInput("Random Command")
        print("Command Input logged")
        
